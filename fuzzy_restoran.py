# ============================================================
# Sistem Fuzzy Logic untuk Pemilihan 5 Restoran Terbaik
# Mata Kuliah: Kecerdasan Buatan
# ============================================================

# ---------- 1. MEMBACA DATA DARI FILE ----------

def baca_data(nama_file):
    """Membaca data restoran dari file xlsx menggunakan openpyxl (bukan library fuzzy)."""
    from openpyxl import load_workbook
    wb = load_workbook(nama_file, read_only=True)
    ws = wb.active

    data = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # Lewati header
        id_resto, pelayanan, harga = row[0], row[1], row[2]
        data.append({
            "id": int(id_resto),
            "pelayanan": float(pelayanan),
            "harga": float(harga)
        })
    return data


# ---------- 2. FUNGSI KEANGGOTAAN INPUT ----------

# --- Kualitas Pelayanan (range: 1 - 100) ---
# Linguistik: Buruk, Sedang, Baik

def pelayanan_buruk(x):
    """Trapesium kiri: turun dari 1 ke 40"""
    if x <= 20:
        return 1.0
    elif x <= 50:
        return (50 - x) / (50 - 20)
    else:
        return 0.0

def pelayanan_sedang(x):
    """Segitiga: naik 30-50, turun 50-70"""
    if x <= 30 or x >= 70:
        return 0.0
    elif x <= 50:
        return (x - 30) / (50 - 30)
    else:
        return (70 - x) / (70 - 50)

def pelayanan_baik(x):
    """Trapesium kanan: naik dari 60 ke 80, tetap 1 setelahnya"""
    if x <= 60:
        return 0.0
    elif x <= 80:
        return (x - 60) / (80 - 60)
    else:
        return 1.0


# --- Harga (range: 25000 - 55000) ---
# Linguistik: Murah, Sedang, Mahal

def harga_murah(x):
    """Trapesium kiri: turun dari 25000 ke 35000"""
    if x <= 28000:
        return 1.0
    elif x <= 37000:
        return (37000 - x) / (37000 - 28000)
    else:
        return 0.0

def harga_sedang(x):
    """Segitiga: naik 32000-42000, turun 42000-52000"""
    if x <= 32000 or x >= 52000:
        return 0.0
    elif x <= 42000:
        return (x - 32000) / (42000 - 32000)
    else:
        return (52000 - x) / (52000 - 42000)

def harga_mahal(x):
    """Trapesium kanan: naik dari 47000 ke 55000"""
    if x <= 47000:
        return 0.0
    elif x <= 55000:
        return (x - 47000) / (55000 - 47000)
    else:
        return 1.0


# ---------- 3. FUNGSI KEANGGOTAAN OUTPUT ----------

# Kelayakan Restoran (range: 0 - 100)
# Linguistik: TidakLayak, CukupLayak, SangatLayak

def output_tidak_layak(z):
    """Trapesium kiri untuk skor output"""
    if z <= 20:
        return 1.0
    elif z <= 40:
        return (40 - z) / (40 - 20)
    else:
        return 0.0

def output_cukup_layak(z):
    """Segitiga tengah"""
    if z <= 30 or z >= 70:
        return 0.0
    elif z <= 50:
        return (z - 30) / (50 - 30)
    else:
        return (70 - z) / (70 - 50)

def output_sangat_layak(z):
    """Trapesium kanan"""
    if z <= 60:
        return 0.0
    elif z <= 80:
        return (z - 60) / (80 - 60)
    else:
        return 1.0


# ---------- 4. FUZZIFIKASI ----------

def fuzzifikasi(pelayanan, harga):
    """Menghitung derajat keanggotaan setiap variabel input."""
    fuzzy = {
        # Pelayanan
        "pelayanan_buruk":  pelayanan_buruk(pelayanan),
        "pelayanan_sedang": pelayanan_sedang(pelayanan),
        "pelayanan_baik":   pelayanan_baik(pelayanan),
        # Harga
        "harga_murah":  harga_murah(harga),
        "harga_sedang": harga_sedang(harga),
        "harga_mahal":  harga_mahal(harga),
    }
    return fuzzy


# ---------- 5. INFERENSI (Mamdani - MIN) ----------

def inferensi(fuzzy):
    """
    Aturan inferensi (9 rules):
    Pelayanan Baik  + Harga Murah  -> Sangat Layak
    Pelayanan Baik  + Harga Sedang -> Sangat Layak
    Pelayanan Baik  + Harga Mahal  -> Cukup Layak
    Pelayanan Sedang + Harga Murah  -> Sangat Layak
    Pelayanan Sedang + Harga Sedang -> Cukup Layak
    Pelayanan Sedang + Harga Mahal  -> Tidak Layak
    Pelayanan Buruk + Harga Murah  -> Cukup Layak
    Pelayanan Buruk + Harga Sedang -> Tidak Layak
    Pelayanan Buruk + Harga Mahal  -> Tidak Layak
    """
    pb  = fuzzy["pelayanan_buruk"]
    ps  = fuzzy["pelayanan_sedang"]
    pba = fuzzy["pelayanan_baik"]
    hm  = fuzzy["harga_murah"]
    hs  = fuzzy["harga_sedang"]
    hma = fuzzy["harga_mahal"]

    # Setiap rule: (alpha, label_output)
    rules = [
        (min(pba, hm),  "sangat_layak"),
        (min(pba, hs),  "sangat_layak"),
        (min(pba, hma), "cukup_layak"),
        (min(ps,  hm),  "sangat_layak"),
        (min(ps,  hs),  "cukup_layak"),
        (min(ps,  hma), "tidak_layak"),
        (min(pb,  hm),  "cukup_layak"),
        (min(pb,  hs),  "tidak_layak"),
        (min(pb,  hma), "tidak_layak"),
    ]

    # Agregasi: ambil nilai alpha maksimum per label output
    agregasi = {
        "tidak_layak":  0.0,
        "cukup_layak":  0.0,
        "sangat_layak": 0.0,
    }
    for alpha, label in rules:
        if alpha > agregasi[label]:
            agregasi[label] = alpha

    return agregasi


# ---------- 6. DEFUZZIFIKASI (Centroid / Center of Area) ----------

def defuzzifikasi(agregasi):
    """
    Metode: Centroid (Center of Area)
    z* = sum(z * mu_agregat(z)) / sum(mu_agregat(z))
    z di-sample dari 0 hingga 100 dengan step 0.5
    """
    step = 0.5
    z_vals = [i * step for i in range(int(100 / step) + 1)]  # 0, 0.5, 1, ..., 100

    numerator   = 0.0
    denominator = 0.0

    for z in z_vals:
        # Hitung derajat keanggotaan output yang sudah di-clip (min dengan alpha)
        mu_tidak   = min(agregasi["tidak_layak"],  output_tidak_layak(z))
        mu_cukup   = min(agregasi["cukup_layak"],  output_cukup_layak(z))
        mu_sangat  = min(agregasi["sangat_layak"], output_sangat_layak(z))

        # Agregasi: ambil maksimum dari semua output
        mu_total = max(mu_tidak, mu_cukup, mu_sangat)

        numerator   += z * mu_total
        denominator += mu_total

    if denominator == 0:
        return 0.0
    return numerator / denominator


# ---------- 7. MENYIMPAN OUTPUT KE FILE ----------

def simpan_output(hasil, nama_file):
    """Menyimpan 5 restoran terbaik ke file xlsx."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Peringkat"

    # Header
    headers = ["Peringkat", "ID Restoran", "Kualitas Pelayanan", "Harga (Rp)", "Skor Kelayakan"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Isi data
    for rank, item in enumerate(hasil, start=1):
        ws.append([
            rank,
            item["id"],
            item["pelayanan"],
            item["harga"],
            round(item["skor"], 4)
        ])

    # Auto-fit kolom
    for col in ws.columns:
        max_len = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    wb.save(nama_file)
    print(f"\nOutput disimpan ke: {nama_file}")


# ---------- 8. PROGRAM UTAMA ----------

def main():
    input_file  = r"C:\Users\MSCRuser\Downloads\restoran.xlsx"
    output_file = r"C:\Users\MSCRuser\Downloads\peringkat.xlsx"

    print("=" * 55)
    print("  SISTEM FUZZY LOGIC - SELEKSI RESTORAN TERBAIK")
    print("=" * 55)

    # Baca data
    data = baca_data(input_file)
    print(f"Jumlah data restoran: {len(data)}")

    # Proses setiap restoran
    hasil = []
    for resto in data:
        fuzzy    = fuzzifikasi(resto["pelayanan"], resto["harga"])
        agregasi = inferensi(fuzzy)
        skor     = defuzzifikasi(agregasi)
        hasil.append({
            "id":        resto["id"],
            "pelayanan": resto["pelayanan"],
            "harga":     resto["harga"],
            "skor":      skor
        })

    # Urutkan berdasarkan skor tertinggi
    hasil.sort(key=lambda x: x["skor"], reverse=True)
    top5 = hasil[:5]

    # Tampilkan hasil
    print("\n--- 5 RESTORAN TERBAIK ---")
    print(f"{'Rank':<6} {'ID':>4} {'Pelayanan':>12} {'Harga (Rp)':>12} {'Skor':>10}")
    print("-" * 50)
    for rank, item in enumerate(top5, start=1):
        print(f"{rank:<6} {item['id']:>4} {item['pelayanan']:>12.0f} {item['harga']:>12.0f} {item['skor']:>10.4f}")

    # Simpan ke file
    simpan_output(top5, output_file)


if __name__ == "__main__":
    main()
